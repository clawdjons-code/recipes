"""Generate Fitness_Tracker.xlsx for upload to Google Drive as a Sheet.

Run: python3 build_tracker.py
Output: Fitness_Tracker.xlsx in the same directory.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "Fitness_Tracker.xlsx"

# Number of pre-filled formula rows on logging tabs.
DAILY_ROWS = 1000
WORKOUT_ROWS = 5000

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(border_style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
GOOD_FONT = Font(color="006100")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
BAD_FONT = Font(color="9C0006")


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_name(wb: Workbook, name: str, ref: str) -> None:
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


# ---------------------------------------------------------------------------
# Workbook setup
# ---------------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------------------
# README tab
# ---------------------------------------------------------------------------
readme = wb.create_sheet("README")
readme["A1"] = "Fitness Tracker — Setup & Usage"
readme["A1"].font = Font(size=16, bold=True)

readme_lines = [
    "",
    "STEP 1 — Upload this .xlsx to Google Drive, then open it and choose",
    "         File → Save as Google Sheets. Work from the Sheets version.",
    "",
    "STEP 2 — Fill in the Config tab (units, starting weight, goal, height,",
    "         age, sex, activity level, target deficit). Everything else",
    "         reads from Config — do not hardcode values anywhere else.",
    "",
    "STEP 3 — Point the Telegram (Hermes) agent at the two append schemas",
    "         below. The agent should append one row per logged event.",
    "",
    "APPEND SCHEMAS (column order matters):",
    "",
    "Daily Log:",
    "  Date | Weight | Kcal In | Kcal Burned (exercise) | <leave rest blank>",
    "  The remaining columns (Net Kcal, Cumulative Deficit, Theoretical Fat",
    "  Lost, Theoretical Weight, Actual Δ, Divergence, 7-day Avg, Notes)",
    "  are filled by formulas. The agent only writes the first 4 + Notes.",
    "",
    "Workout Log:",
    "  Date | Exercise | Set | Reps | Weight | RPE | Notes",
    "  The Volume, e1RM, and Is PR columns are formula-driven.",
    "",
    "TABS:",
    "  Config             — your settings; edit these.",
    "  Daily Log          — daily weigh-ins + calories. Agent appends here.",
    "  Workout Log        — every set you do. Agent appends here.",
    "  Exercise Library   — list of exercises, auto-updates PR & last session.",
    "  PRs                — best weight + best e1RM per exercise.",
    "  Body Measurements  — optional weekly measurements.",
    "  Dashboard          — charts.",
    "  Exercise Lookup    — pick an exercise from the dropdown to see your",
    "                       full history and a progression chart.",
    "",
    "FORMULAS USED (work in both Excel 365 and Google Sheets):",
    "  Mifflin-St Jeor BMR, FILTER, XLOOKUP, MAXIFS, AVERAGEIFS, TEXTJOIN.",
    "",
    "THEORETICAL FAT LOSS:",
    "  cumulative deficit ÷ (3500 kcal/lb or 7700 kcal/kg, per Config units).",
    "  This is a model, not a measurement. Reality lags or leads it; the",
    "  Divergence column shows the gap.",
]
for i, line in enumerate(readme_lines, start=2):
    readme[f"A{i}"] = line
autosize(readme, {"A": 90})

# ---------------------------------------------------------------------------
# Config tab
# ---------------------------------------------------------------------------
cfg = wb.create_sheet("Config")
cfg["A1"] = "Setting"
cfg["B1"] = "Value"
cfg["C1"] = "Notes"
style_header(cfg, 1, 3)

config_rows = [
    ("Units",                "lbs/in",   "Type 'lbs/in' or 'kg/cm'"),
    ("Sex",                  "M",        "'M' or 'F' (Mifflin-St Jeor)"),
    ("Age",                  30,         "years"),
    ("Height",               70,         "inches if lbs/in, cm if kg/cm"),
    ("Starting Weight",      200,        "lbs or kg, per Units"),
    ("Starting Date",        date.today(), "first day you're tracking"),
    ("Goal Weight",          175,        "lbs or kg"),
    ("Activity Level",       "Moderate", "Sedentary / Light / Moderate / Active / Very Active"),
    ("Activity Multiplier",  '=IFERROR(VLOOKUP(B9,{"Sedentary",1.2;"Light",1.375;"Moderate",1.55;"Active",1.725;"Very Active",1.9},2,FALSE),1.55)', "auto from Activity Level"),
    ("Weight (kg)",          '=IF(B2="kg/cm",B6,B6/2.2046226218)', "auto"),
    ("Height (cm)",          '=IF(B2="kg/cm",B5,B5*2.54)',         "auto"),
    ("BMR (kcal/day)",       '=IF(B3="M",10*B11+6.25*B12-5*B4+5,10*B11+6.25*B12-5*B4-161)', "Mifflin-St Jeor"),
    ("TDEE (kcal/day)",      '=B13*B10', "BMR × activity"),
    ("Target Daily Deficit", 500,        "kcal below TDEE"),
    ("Kcal per lb fat",      3500,       "constant"),
    ("Kcal per kg fat",      7700,       "constant"),
    ("Fat constant (active)",'=IF(B2="kg/cm",B17,B16)', "auto"),
]
for i, (label, value, note) in enumerate(config_rows, start=2):
    cfg.cell(row=i, column=1, value=label).font = Font(bold=True)
    cfg.cell(row=i, column=2, value=value)
    cfg.cell(row=i, column=3, value=note).font = Font(italic=True, color="666666")
    for col in (1, 2, 3):
        cfg.cell(row=i, column=col).border = BORDER

cfg["B7"].number_format = "yyyy-mm-dd"
autosize(cfg, {"A": 24, "B": 18, "C": 50})

# Named ranges — reference Config from anywhere
add_name(wb, "Cfg_Units",         "Config!$B$2")
add_name(wb, "Cfg_StartWeight",   "Config!$B$6")
add_name(wb, "Cfg_StartDate",     "Config!$B$7")
add_name(wb, "Cfg_GoalWeight",    "Config!$B$8")
add_name(wb, "Cfg_TDEE",          "Config!$B$14")
add_name(wb, "Cfg_TargetDeficit", "Config!$B$15")
add_name(wb, "Cfg_FatConstant",   "Config!$B$18")


# ---------------------------------------------------------------------------
# Daily Log tab
# ---------------------------------------------------------------------------
daily = wb.create_sheet("Daily Log")
daily_headers = [
    "Date", "Weight", "Kcal In", "Kcal Burned",
    "Net Kcal", "Cumulative Deficit", "Theoretical Fat Lost",
    "Theoretical Weight", "Actual Δ", "Divergence",
    "7-day Avg Weight", "Notes",
]
for i, h in enumerate(daily_headers, start=1):
    daily.cell(row=1, column=i, value=h)
style_header(daily, 1, len(daily_headers))

for r in range(2, DAILY_ROWS + 2):
    # E: Net Kcal = (Kcal In) - TDEE - (Kcal Burned)
    daily[f"E{r}"] = f'=IF(C{r}="","",C{r}-Cfg_TDEE-IFERROR(D{r},0))'
    # F: Cumulative Deficit (running, positive = deficit). Assumes rows in date order.
    if r == 2:
        daily[f"F{r}"] = f'=IF(E{r}="","",-E{r})'
    else:
        daily[f"F{r}"] = f'=IF(E{r}="","",IFERROR(F{r-1},0)+(-E{r}))'
    # G: Theoretical Fat Lost
    daily[f"G{r}"] = f'=IF(F{r}="","",F{r}/Cfg_FatConstant)'
    # H: Theoretical Weight
    daily[f"H{r}"] = f'=IF(G{r}="","",Cfg_StartWeight-G{r})'
    # I: Actual Δ (positive = lost weight)
    daily[f"I{r}"] = f'=IF(B{r}="","",Cfg_StartWeight-B{r})'
    # J: Divergence (positive = ahead of plan)
    daily[f"J{r}"] = f'=IF(OR(B{r}="",G{r}=""),"",I{r}-G{r})'
    # K: 7-day avg weight
    daily[f"K{r}"] = (
        f'=IF(B{r}="","",'
        f'IFERROR(AVERAGEIFS($B$2:$B${DAILY_ROWS+1},'
        f'$A$2:$A${DAILY_ROWS+1},">="&A{r}-6,'
        f'$A$2:$A${DAILY_ROWS+1},"<="&A{r}),""))'
    )

# Number formats
for r in range(2, DAILY_ROWS + 2):
    daily[f"A{r}"].number_format = "yyyy-mm-dd"
    daily[f"B{r}"].number_format = "0.0"
    daily[f"G{r}"].number_format = "0.00"
    daily[f"H{r}"].number_format = "0.0"
    daily[f"I{r}"].number_format = "0.0"
    daily[f"J{r}"].number_format = "+0.00;-0.00;0"
    daily[f"K{r}"].number_format = "0.00"

# Conditional formatting on Divergence (column J)
daily.conditional_formatting.add(
    f"J2:J{DAILY_ROWS+1}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=GOOD_FILL, font=GOOD_FONT),
)
daily.conditional_formatting.add(
    f"J2:J{DAILY_ROWS+1}",
    CellIsRule(operator="lessThan", formula=["0"], fill=BAD_FILL, font=BAD_FONT),
)

autosize(daily, {
    "A": 12, "B": 9, "C": 9, "D": 12, "E": 11, "F": 16,
    "G": 18, "H": 16, "I": 11, "J": 12, "K": 16, "L": 30,
})
daily.freeze_panes = "A2"

add_name(wb, "DL_Date",   f"'Daily Log'!$A$2:$A${DAILY_ROWS+1}")
add_name(wb, "DL_Weight", f"'Daily Log'!$B$2:$B${DAILY_ROWS+1}")


# ---------------------------------------------------------------------------
# Workout Log tab
# ---------------------------------------------------------------------------
wlog = wb.create_sheet("Workout Log")
wlog_headers = ["Date", "Exercise", "Set", "Reps", "Weight", "RPE", "Notes",
                "Volume", "e1RM", "Is PR"]
for i, h in enumerate(wlog_headers, start=1):
    wlog.cell(row=1, column=i, value=h)
style_header(wlog, 1, len(wlog_headers))

for r in range(2, WORKOUT_ROWS + 2):
    # H Volume = Reps * Weight
    wlog[f"H{r}"] = f'=IF(OR(D{r}="",E{r}=""),"",D{r}*E{r})'
    # I e1RM (Epley)
    wlog[f"I{r}"] = f'=IF(OR(D{r}="",E{r}=""),"",E{r}*(1+D{r}/30))'
    # J Is PR: this row's e1RM ≥ max e1RM for this exercise on or before this date
    wlog[f"J{r}"] = (
        f'=IF(I{r}="","",'
        f'IF(I{r}>=MAXIFS($I$2:$I${WORKOUT_ROWS+1},'
        f'$B$2:$B${WORKOUT_ROWS+1},B{r},'
        f'$A$2:$A${WORKOUT_ROWS+1},"<="&A{r}),"PR",""))'
    )

for r in range(2, WORKOUT_ROWS + 2):
    wlog[f"A{r}"].number_format = "yyyy-mm-dd"
    wlog[f"H{r}"].number_format = "0"
    wlog[f"I{r}"].number_format = "0.0"

# Highlight PR rows
wlog.conditional_formatting.add(
    f"A2:J{WORKOUT_ROWS+1}",
    FormulaRule(formula=[f'$J2="PR"'], fill=GOOD_FILL, font=GOOD_FONT),
)

autosize(wlog, {
    "A": 12, "B": 22, "C": 5, "D": 6, "E": 8, "F": 5,
    "G": 30, "H": 9, "I": 9, "J": 6,
})
wlog.freeze_panes = "A2"

add_name(wb, "WL_Date",     f"'Workout Log'!$A$2:$A${WORKOUT_ROWS+1}")
add_name(wb, "WL_Exercise", f"'Workout Log'!$B$2:$B${WORKOUT_ROWS+1}")
add_name(wb, "WL_Set",      f"'Workout Log'!$C$2:$C${WORKOUT_ROWS+1}")
add_name(wb, "WL_Reps",     f"'Workout Log'!$D$2:$D${WORKOUT_ROWS+1}")
add_name(wb, "WL_Weight",   f"'Workout Log'!$E$2:$E${WORKOUT_ROWS+1}")
add_name(wb, "WL_RPE",      f"'Workout Log'!$F$2:$F${WORKOUT_ROWS+1}")
add_name(wb, "WL_Volume",   f"'Workout Log'!$H$2:$H${WORKOUT_ROWS+1}")
add_name(wb, "WL_e1RM",     f"'Workout Log'!$I$2:$I${WORKOUT_ROWS+1}")


# ---------------------------------------------------------------------------
# Exercise Library tab
# ---------------------------------------------------------------------------
lib = wb.create_sheet("Exercise Library")
lib_headers = [
    "Exercise", "Category", "Primary Muscle", "Secondary", "Equipment",
    "Heaviest Weight", "Best e1RM", "Last Performed", "Last Session",
]
for i, h in enumerate(lib_headers, start=1):
    lib.cell(row=1, column=i, value=h)
style_header(lib, 1, len(lib_headers))

# Seed common exercises (Exercise, Category, Primary, Secondary, Equipment)
seed_exercises = [
    ("Back Squat",       "Legs",  "Quads",      "Glutes, Hamstrings", "Barbell"),
    ("Front Squat",      "Legs",  "Quads",      "Glutes, Core",       "Barbell"),
    ("Deadlift",         "Pull",  "Posterior",  "Back, Hamstrings",   "Barbell"),
    ("Romanian Deadlift","Pull",  "Hamstrings", "Glutes, Back",       "Barbell"),
    ("Bench Press",      "Push",  "Chest",      "Triceps, Shoulders", "Barbell"),
    ("Incline Bench",    "Push",  "Upper Chest","Triceps, Shoulders", "Barbell"),
    ("Overhead Press",   "Push",  "Shoulders",  "Triceps",            "Barbell"),
    ("Barbell Row",      "Pull",  "Back",       "Biceps, Rear Delts", "Barbell"),
    ("Pull-Up",          "Pull",  "Lats",       "Biceps, Back",       "Bodyweight"),
    ("Chin-Up",          "Pull",  "Biceps",     "Lats",               "Bodyweight"),
    ("Dip",              "Push",  "Chest",      "Triceps",            "Bodyweight"),
    ("Push-Up",          "Push",  "Chest",      "Triceps, Shoulders", "Bodyweight"),
    ("Decline Push-Up",  "Push",  "Upper Chest","Triceps, Shoulders", "Bodyweight"),
    ("Incline Push-Up",  "Push",  "Lower Chest","Triceps, Shoulders", "Bodyweight"),
    ("Diamond Push-Up",  "Push",  "Triceps",    "Chest, Shoulders",   "Bodyweight"),
    ("Lunge",            "Legs",  "Quads",      "Glutes",             "Dumbbell"),
    ("Bulgarian Split",  "Legs",  "Quads",      "Glutes",             "Dumbbell"),
    ("Hip Thrust",       "Legs",  "Glutes",     "Hamstrings",         "Barbell"),
    ("Dumbbell Row",     "Pull",  "Back",       "Biceps",             "Dumbbell"),
    ("DB Bench",         "Push",  "Chest",      "Triceps, Shoulders", "Dumbbell"),
    ("DB Press",         "Push",  "Shoulders",  "Triceps",            "Dumbbell"),
    ("Lateral Raise",    "Push",  "Side Delts", "",                   "Dumbbell"),
    ("Curl",             "Pull",  "Biceps",     "Forearms",           "Dumbbell"),
    ("Tricep Extension", "Push",  "Triceps",    "",                   "Dumbbell"),
    ("Plank",            "Core",  "Core",       "",                   "Bodyweight"),
    ("Hanging Leg Raise","Core",  "Abs",        "Hip Flexors",        "Bodyweight"),
    ("Run",              "Cardio","Legs",       "",                   "Cardio"),
    ("Bike",             "Cardio","Legs",       "",                   "Cardio"),
    ("Row (Erg)",        "Cardio","Full Body",  "",                   "Cardio"),
]
LIB_ROWS = 200  # extra rows for user-added exercises
for i, row in enumerate(seed_exercises, start=2):
    for j, val in enumerate(row, start=1):
        lib.cell(row=i, column=j, value=val)

# Pre-fill formula columns for full LIB_ROWS range
for r in range(2, LIB_ROWS + 2):
    lib[f"F{r}"] = f'=IF(A{r}="","",IFERROR(MAXIFS(WL_Weight,WL_Exercise,A{r}),""))'
    lib[f"G{r}"] = f'=IF(A{r}="","",IFERROR(MAXIFS(WL_e1RM,WL_Exercise,A{r}),""))'
    lib[f"H{r}"] = f'=IF(A{r}="","",IFERROR(MAXIFS(WL_Date,WL_Exercise,A{r}),""))'
    # Last session: concat all sets from the most-recent date
    lib[f"I{r}"] = (
        f'=IF(OR(A{r}="",H{r}=""),"",'
        f'IFERROR(TEXTJOIN(", ",TRUE,'
        f'IF((WL_Exercise=A{r})*(WL_Date=H{r}),WL_Reps&"×"&WL_Weight,"")),""))'
    )
    lib[f"H{r}"].number_format = "yyyy-mm-dd"
    lib[f"F{r}"].number_format = "0.0"
    lib[f"G{r}"].number_format = "0.0"

autosize(lib, {
    "A": 22, "B": 10, "C": 16, "D": 22, "E": 12,
    "F": 14, "G": 12, "H": 14, "I": 36,
})
lib.freeze_panes = "A2"

add_name(wb, "Lib_Exercise", f"'Exercise Library'!$A$2:$A${LIB_ROWS+1}")


# Dropdown validation: Exercise column on Workout Log uses Library names
dv = DataValidation(
    type="list",
    formula1=f"='Exercise Library'!$A$2:$A${LIB_ROWS+1}",
    allow_blank=True,
)
dv.error = "Exercise not in library. Add it to Exercise Library first."
dv.errorTitle = "Unknown exercise"
dv.prompt = "Pick from Exercise Library"
dv.promptTitle = "Exercise"
wlog.add_data_validation(dv)
dv.add(f"B2:B{WORKOUT_ROWS+1}")


# ---------------------------------------------------------------------------
# PRs tab
# ---------------------------------------------------------------------------
prs = wb.create_sheet("PRs")
pr_headers = [
    "Exercise", "Heaviest Weight", "Heaviest Date",
    "Best e1RM", "Best e1RM Date", "Best e1RM Reps", "Best e1RM Weight",
]
for i, h in enumerate(pr_headers, start=1):
    prs.cell(row=1, column=i, value=h)
style_header(prs, 1, len(pr_headers))

for r in range(2, LIB_ROWS + 2):
    # Mirror exercise list from Library
    prs[f"A{r}"] = f"='Exercise Library'!A{r}"
    # Heaviest weight + its date
    prs[f"B{r}"] = f'=IF(A{r}="","",IFERROR(MAXIFS(WL_Weight,WL_Exercise,A{r}),""))'
    prs[f"C{r}"] = (
        f'=IF(OR(A{r}="",B{r}=""),"",'
        f'IFERROR(XLOOKUP(1,(WL_Exercise=A{r})*(WL_Weight=B{r}),WL_Date,"",0),""))'
    )
    # Best e1RM + date + the reps/weight that produced it
    prs[f"D{r}"] = f'=IF(A{r}="","",IFERROR(MAXIFS(WL_e1RM,WL_Exercise,A{r}),""))'
    prs[f"E{r}"] = (
        f'=IF(OR(A{r}="",D{r}=""),"",'
        f'IFERROR(XLOOKUP(1,(WL_Exercise=A{r})*(WL_e1RM=D{r}),WL_Date,"",0),""))'
    )
    prs[f"F{r}"] = (
        f'=IF(OR(A{r}="",D{r}=""),"",'
        f'IFERROR(XLOOKUP(1,(WL_Exercise=A{r})*(WL_e1RM=D{r}),WL_Reps,"",0),""))'
    )
    prs[f"G{r}"] = (
        f'=IF(OR(A{r}="",D{r}=""),"",'
        f'IFERROR(XLOOKUP(1,(WL_Exercise=A{r})*(WL_e1RM=D{r}),WL_Weight,"",0),""))'
    )
    prs[f"C{r}"].number_format = "yyyy-mm-dd"
    prs[f"E{r}"].number_format = "yyyy-mm-dd"
    prs[f"B{r}"].number_format = "0.0"
    prs[f"D{r}"].number_format = "0.0"
    prs[f"G{r}"].number_format = "0.0"

autosize(prs, {"A": 22, "B": 16, "C": 14, "D": 12, "E": 14, "F": 16, "G": 18})
prs.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Body Measurements tab
# ---------------------------------------------------------------------------
body = wb.create_sheet("Body Measurements")
body_headers = ["Date", "Waist", "Chest", "Arms", "Thighs", "Hips", "Body Fat %", "Notes"]
for i, h in enumerate(body_headers, start=1):
    body.cell(row=1, column=i, value=h)
style_header(body, 1, len(body_headers))
for r in range(2, 200):
    body[f"A{r}"].number_format = "yyyy-mm-dd"
    body[f"G{r}"].number_format = "0.0%"
autosize(body, {"A": 12, "B": 8, "C": 8, "D": 8, "E": 8, "F": 8, "G": 12, "H": 30})
body.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Dashboard tab
# ---------------------------------------------------------------------------
dash = wb.create_sheet("Dashboard")
dash["A1"] = "Dashboard"
dash["A1"].font = Font(size=16, bold=True)
dash["A2"] = "Charts pull live from Daily Log and Workout Log. Add data to see lines populate."
dash["A2"].font = Font(italic=True, color="666666")

# Chart 1: Weight — actual, 7-day avg, theoretical
chart1 = LineChart()
chart1.title = "Weight: Actual vs Theoretical"
chart1.y_axis.title = "Weight"
chart1.x_axis.title = "Date"
chart1.height = 10
chart1.width = 20
dates = Reference(daily, min_col=1, min_row=2, max_row=DAILY_ROWS + 1)
weight_actual = Reference(daily, min_col=2, min_row=1, max_row=DAILY_ROWS + 1)
weight_theoretical = Reference(daily, min_col=8, min_row=1, max_row=DAILY_ROWS + 1)
weight_avg = Reference(daily, min_col=11, min_row=1, max_row=DAILY_ROWS + 1)
chart1.add_data(weight_actual, titles_from_data=True)
chart1.add_data(weight_theoretical, titles_from_data=True)
chart1.add_data(weight_avg, titles_from_data=True)
chart1.set_categories(dates)
dash.add_chart(chart1, "A4")

# Chart 2: Cumulative deficit
chart2 = LineChart()
chart2.title = "Cumulative Calorie Deficit"
chart2.y_axis.title = "kcal"
chart2.x_axis.title = "Date"
chart2.height = 10
chart2.width = 20
deficit = Reference(daily, min_col=6, min_row=1, max_row=DAILY_ROWS + 1)
chart2.add_data(deficit, titles_from_data=True)
chart2.set_categories(dates)
dash.add_chart(chart2, "A26")

# Chart 3: Daily kcal in
chart3 = LineChart()
chart3.title = "Daily Calories In"
chart3.y_axis.title = "kcal"
chart3.x_axis.title = "Date"
chart3.height = 10
chart3.width = 20
kcal_in = Reference(daily, min_col=3, min_row=1, max_row=DAILY_ROWS + 1)
chart3.add_data(kcal_in, titles_from_data=True)
chart3.set_categories(dates)
dash.add_chart(chart3, "A48")

# Chart 4: Daily workout volume (sum across sets per day)
# Helper columns on Dashboard out of the way
dash["P1"] = "Date"
dash["Q1"] = "Total Volume"
dash["P1"].font = Font(bold=True)
dash["Q1"].font = Font(bold=True)
# Build a 90-day rolling window of dates and volumes
for i in range(2, 92):
    offset = 91 - i  # i=2 => 89 days ago, i=91 => today
    dash[f"P{i}"] = f"=TODAY()-{offset}"
    dash[f"Q{i}"] = f'=IFERROR(SUMIFS(WL_Volume,WL_Date,P{i}),0)'
    dash[f"P{i}"].number_format = "yyyy-mm-dd"
chart4 = BarChart()
chart4.title = "Workout Volume (last 90 days)"
chart4.y_axis.title = "Total Volume (reps × weight)"
chart4.x_axis.title = "Date"
chart4.height = 10
chart4.width = 20
vol_dates = Reference(dash, min_col=16, min_row=2, max_row=91)
vol_values = Reference(dash, min_col=17, min_row=1, max_row=91)
chart4.add_data(vol_values, titles_from_data=True)
chart4.set_categories(vol_dates)
dash.add_chart(chart4, "A70")

autosize(dash, {"A": 12, "P": 12, "Q": 14})


# ---------------------------------------------------------------------------
# Exercise Lookup tab
# ---------------------------------------------------------------------------
look = wb.create_sheet("Exercise Lookup")
look["A1"] = "Pick exercise:"
look["A1"].font = Font(bold=True)
look["B1"] = "Decline Push-Up"  # default selection
look["B1"].fill = PatternFill("solid", fgColor="FFF2CC")
look["B1"].font = Font(bold=True, size=14)
look["B1"].alignment = Alignment(horizontal="center")
look["B1"].border = BORDER

# Dropdown for B1
dv_lookup = DataValidation(
    type="list",
    formula1=f"='Exercise Library'!$A$2:$A${LIB_ROWS+1}",
    allow_blank=False,
)
look.add_data_validation(dv_lookup)
dv_lookup.add("B1")

# Quick summary row
look["A3"] = "Last performed:"
look["B3"] = '=IFERROR(MAXIFS(WL_Date,WL_Exercise,B1),"never")'
look["B3"].number_format = "yyyy-mm-dd"
look["A4"] = "Last session:"
look["B4"] = (
    '=IFERROR(TEXTJOIN(", ",TRUE,'
    'IF((WL_Exercise=B1)*(WL_Date=B3),WL_Reps&"×"&WL_Weight,"")),"")'
)
look["A5"] = "Heaviest ever:"
look["B5"] = '=IFERROR(MAXIFS(WL_Weight,WL_Exercise,B1),"")'
look["A6"] = "Best e1RM:"
look["B6"] = '=IFERROR(MAXIFS(WL_e1RM,WL_Exercise,B1),"")'
for r in (3, 4, 5, 6):
    look.cell(row=r, column=1).font = Font(bold=True)

# Filtered history (spilled). Modern Excel & Sheets render this as a dynamic array.
look["A8"] = "Full history (most recent first):"
look["A8"].font = Font(bold=True)
look["A9"] = "Date"
look["B9"] = "Set"
look["C9"] = "Reps"
look["D9"] = "Weight"
look["E9"] = "RPE"
look["F9"] = "e1RM"
look["G9"] = "Notes"
style_header(look, 9, 7)
look["A10"] = (
    '=IFERROR(SORT(FILTER(CHOOSECOLS('
    "'Workout Log'!A2:I"+str(WORKOUT_ROWS+1)+",1,3,4,5,6,9,7),"
    "WL_Exercise=B1),1,-1),\"No sets logged yet for this exercise.\")"
)

# Chart helper: per-date top-set weight + e1RM (uses helper columns J:L on this tab)
look["J9"] = "Date"
look["K9"] = "Top-Set Weight"
look["L9"] = "Best e1RM (that day)"
style_header(look, 9, 12)
# Build a list of unique dates that have data for selected exercise (last 60 sessions max)
for i in range(10, 70):
    # i-th most-recent date for B1
    look[f"J{i}"] = (
        f'=IFERROR(LARGE(IF(WL_Exercise=B1,WL_Date),{i-9}),"")'
    )
    look[f"K{i}"] = (
        f'=IF(J{i}="","",IFERROR(MAXIFS(WL_Weight,WL_Exercise,B1,WL_Date,J{i}),""))'
    )
    look[f"L{i}"] = (
        f'=IF(J{i}="","",IFERROR(MAXIFS(WL_e1RM,WL_Exercise,B1,WL_Date,J{i}),""))'
    )
    look[f"J{i}"].number_format = "yyyy-mm-dd"
    look[f"K{i}"].number_format = "0.0"
    look[f"L{i}"].number_format = "0.0"

# Chart for selected exercise progression
chart_lk = LineChart()
chart_lk.title = "Progression (selected exercise)"
chart_lk.y_axis.title = "Weight / e1RM"
chart_lk.x_axis.title = "Date"
chart_lk.height = 10
chart_lk.width = 18
lk_dates = Reference(look, min_col=10, min_row=10, max_row=69)
lk_top = Reference(look, min_col=11, min_row=9, max_row=69)
lk_e1rm = Reference(look, min_col=12, min_row=9, max_row=69)
chart_lk.add_data(lk_top, titles_from_data=True)
chart_lk.add_data(lk_e1rm, titles_from_data=True)
chart_lk.set_categories(lk_dates)
look.add_chart(chart_lk, "I1")

autosize(look, {
    "A": 12, "B": 6, "C": 6, "D": 8, "E": 5, "F": 8, "G": 28,
    "J": 12, "K": 14, "L": 18,
})
look.freeze_panes = "A10"


# ---------------------------------------------------------------------------
# Tab order
# ---------------------------------------------------------------------------
wb._sheets = [
    readme, cfg, daily, wlog, lib, prs, body, dash, look,
]

wb.save(OUT)
print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")
